import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import asyncio
import threading
from upload_script import run_upload
from openpyxl import load_workbook



def clean_excel_spaces(excel_path, log_box):
    from datetime import datetime
    wb = load_workbook(excel_path)
    sheet = wb.active
    count = 0
    for row in sheet.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and any(x in val for x in ["/", "-"]):
                cleaned = val.strip()
                if cleaned != val:
                    log_box.insert(tk.END, f"🧹 Cleaned date: '{val}' → '{cleaned}'\n")
                    cell.value = cleaned
                    count += 1

    wb.save(excel_path)

    log_box.insert(tk.END, f"Cleaned {count} cells.\n\n")



def run_async(loop, log_box):
    asyncio.set_event_loop(loop)
    loop.run_until_complete(run_upload())
    log_box.insert(tk.END, "Upload process completed!\n")



# Tkinter UI


def main_ui():
    root = tk.Tk()

    root.title(" Bullseye Uploader")
    root.geometry("650x500")
    root.config(bg="#1e1e1e")

    #Title

    tk.Label(root, text="Bullseye Automation Panel", fg="#00ffae", bg="#1e1e1e",
             font=("Segoe UI", 16, "bold")).pack(pady=10)
    # Fileselector

    file_path_var = tk.StringVar()
    def browse_file():

        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

        if path:
            file_path_var.set(path)

            log_box.insert(tk.END, f" Selected: {path}\n")
    tk.Button(root, text=" Select Excel File", command=browse_file,

              bg="#0078d7", fg="white", font=("Segoe UI", 10, "bold")).pack(pady=10)

    tk.Entry(root, textvariable=file_path_var, width=60).pack(pady=5)
    # Log Box

    log_box = scrolledtext.ScrolledText(root, width=75, height=20, bg="#121212",

                                        fg="#d4d4d4", insertbackground="white")
    log_box.pack(pady=10)

    # Buttons

    button_frame = tk.Frame(root, bg="#1e1e1e")

    button_frame.pack(pady=10)

    def clean_file():
        path = file_path_var.get()

        if not path:
            messagebox.showwarning("Missing File", "Please select an Excel file first.")

            return
        clean_excel_spaces(path, log_box)

    def start_upload():

        log_box.insert(tk.END, "Starting upload..\n")

        loop = asyncio.new_event_loop()

        t = threading.Thread(target=run_async, args=(loop, log_box))

        t.start()

    tk.Button(button_frame, text="Clean Dates", command=clean_file,

              bg="#ffaa00", fg="black", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=10)

    tk.Button(button_frame, text="Run Upload", command=start_upload,

              bg="#28a745", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=1, padx=10)

    tk.Button(button_frame, text="Exit", command=root.destroy,

              bg="#d9534f", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=2, padx=10)

    root.mainloop()


if __name__ == "__main__":

    main_ui()
 